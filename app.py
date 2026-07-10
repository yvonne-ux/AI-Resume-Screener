import os
import io
import json
import re
import signal
import sqlite3
import time
from datetime import datetime
import anthropic
import pdfplumber
import openpyxl
from docx import Document
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename


class _ClaudeTimeout(Exception):
    pass

def _alarm_handler(signum, frame):
    raise _ClaudeTimeout()


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB to support 20+ resumes

# ── Pipeline database ──
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pipeline.db')

def init_db():
    conn = sqlite3.connect(DATABASE)
    conn.execute('''CREATE TABLE IF NOT EXISTS pipeline (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_name TEXT NOT NULL,
        filename       TEXT,
        role_title     TEXT,
        score          INTEGER,
        stage          TEXT DEFAULT 'submitted',
        key_skills     TEXT DEFAULT '[]',
        strengths      TEXT DEFAULT '[]',
        years_experience INTEGER,
        created_at     TEXT,
        updated_at     TEXT
    )''')
    conn.commit()
    conn.close()

try:
    init_db()
except Exception as e:
    print(f"Warning: Could not initialise pipeline database: {e}")

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def get_placed_examples(role_title, limit=3):
    """Fetch candidates previously placed in similar roles for few-shot learning."""
    try:
        conn = get_db()
        keywords = [kw.strip() for kw in role_title.lower().split() if len(kw) > 2]
        if not keywords:
            conn.close()
            return []
        conditions = ' OR '.join(['LOWER(role_title) LIKE ?' for _ in keywords])
        params = [f'%{kw}%' for kw in keywords] + [limit]
        rows = conn.execute(
            f'''SELECT candidate_name, key_skills, strengths, years_experience, stage
                FROM pipeline
                WHERE stage IN ('offer','interview') AND ({conditions})
                ORDER BY CASE stage WHEN 'offer' THEN 1 ELSE 2 END, updated_at DESC
                LIMIT ?''', params
        ).fetchall()
        conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        print(f"Warning: Could not fetch pipeline examples: {e}")
        return []

ALLOWED_EXTENSIONS = {"pdf", "doc", "docx"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def clean_text(text):
    text = text.replace("\u00a0", "\n").replace("\u2028", "\n")
    text = text.replace("\u00ad", " ").replace("\ufeff", "")
    text = text.encode("utf-8", errors="replace").decode("utf-8")
    return text.strip()


def extract_text_from_pdf(file_bytes):
    pages = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                pages.append(clean_text(page_text))
    return "\n".join(pages)


def extract_text_from_docx(file_bytes):
    """Extract text from DOCX paragraphs."""
    t0 = time.time()
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    print(f"[DOCX] parsed {len(file_bytes)//1024}KB in {time.time()-t0:.2f}s → {len(paragraphs)} paragraphs")
    return clean_text('\n'.join(paragraphs))


def extract_text(file_bytes, filename):
    ext = filename.rsplit(".", 1)[1].lower()
    if ext == "pdf":
        return extract_text_from_pdf(file_bytes)
    elif ext in ("doc", "docx"):
        return extract_text_from_docx(file_bytes)
    return ""


def analyze_jd_with_claude(job_title, job_description):
    client = anthropic.Anthropic(
        api_key=os.environ.get("ANTHROPIC_API_KEY"),
        timeout=50.0,
    )

    prompt = f"""You are an expert recruiter. Analyze this job description and extract what a junior recruiter needs to know to screen candidates effectively.

Job Title: {job_title}
Job Description:
{job_description[:6000]}

Respond with ONLY a valid JSON object (no markdown, no explanation):
{{
  "role_summary": "2-3 sentence plain-English explanation of what this role does and what kind of person would succeed in it",
  "must_have": ["Critical requirement 1", "Critical requirement 2", "Critical requirement 3"],
  "nice_to_have": ["Good-to-have 1", "Good-to-have 2"],
  "watch_out_for": ["Common mismatch or red flag to watch for 1", "Red flag 2"]
}}"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=512,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = message.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


def score_resume_with_claude(cv_text, job_title, job_description, examples=None):
    client = anthropic.Anthropic(
        api_key=os.environ.get("ANTHROPIC_API_KEY"),
        timeout=50.0,   # fail fast — Railway proxy cuts at 60s
    )

    # Build few-shot examples from past placements
    examples_section = ""
    if examples:
        examples_section = "\n\nDHC PAST PLACEMENTS — Real candidates DHC successfully placed in this type of role. Use as scoring benchmarks:\n"
        for ex in examples:
            skills = ex.get('key_skills', '[]')
            if isinstance(skills, str):
                try: skills = json.loads(skills)
                except: skills = []
            tag = "✅ Placed via Offer" if ex.get('stage') == 'offer' else "📞 Placed via Interview"
            examples_section += f"- {ex.get('candidate_name','?')} | {tag} | Skills: {', '.join(skills[:5])} | {ex.get('years_experience','?')} yrs exp\n"

    prompt = f"""You are a senior healthcare/professional services recruiter with 15 years of experience. Evaluate this CV fairly and accurately against the job requirements below.
{examples_section}
Job Title: {job_title}
Job Description:
{job_description[:4000]}

CV:
---
{cv_text[:6000]}
---

IMPORTANT EVALUATION RULES:

1. SENIORITY TITLE INFLATION — NEVER penalise for title mismatch alone. Job titles mean different things across industries and companies. Always evaluate based on actual scope of work, responsibilities, and team size — NOT the job title. Examples of common title inflation to be aware of:
   - Banking/Finance: "Vice President (VP)" is often equivalent to a Manager or Senior Executive in other sectors
   - MNCs: "Director" can mean an individual contributor with no direct reports
   - Startups: "Head of" or "Lead" may mean a team of 1-2 people
   - Government/Statutory Boards: grades like "MX13" or "Band 3" indicate seniority, not the title
   - When a candidate's title appears senior but their scope, team size, and responsibilities match the role, do NOT flag as overqualified
   - Only flag overqualified=true if both the title AND the described scope/responsibilities are clearly above the role requirements

2. TRANSFERABLE EXPERIENCE: Give full credit for transferable skills across sectors. Examples:
   - Government/Statutory Board experience (HPB, MOH, CPF, HDB, MOE, VITAL) = strong operations, policy, and governance skills
   - Military (SAF/SPF/SCDF) backgrounds = strong operations management, project delivery, discipline
   - Big 4 consulting (Deloitte, KPMG, PwC, EY, Accenture) = structured methodology and stakeholder management
   - Healthcare-adjacent bodies (HPB, AIC, NCSS, VWOs) count as healthcare experience
   - Banking/finance ops experience transfers well to healthcare admin and HR shared services roles
   - PSA (Patient Service Associate) roles DO NOT require prior hospital or healthcare experience — any strong customer service, counter service, admin, operations, or logistics background is sufficient and should be scored positively
   - Hospitality & attractions (hotels, theme parks like Resorts World Sentosa, Universal Studios, airlines, cruise lines) = strong customer service, high-volume people management, safety protocols, emergency response — directly transferable to PSA roles
   - Retail and F&B experience = customer service, cash handling, complaint resolution — transferable to PSA, clinic admin, and front-desk healthcare roles
   - Call centre / BPO experience = high-volume customer interaction, scripted communication, escalation handling — transferable to PSA and patient-facing roles
   - Operations, logistics, admin, and counter service from ANY industry (government, community, sports, retail, transport) = valid and strong background for PSA roles

3. SINGAPORE ACRONYMS & ORGANISATIONS — Recognise these as strong signals, not unknown text:
   MOH=Ministry of Health, NHG=National Healthcare Group, SingHealth=Singapore Health Services, NUH=National University Hospital, TTSH=Tan Tock Seng Hospital, SGH=Singapore General Hospital, KTPH=Khoo Teck Puat Hospital, HPB=Health Promotion Board, AIC=Agency for Integrated Care, VITAL=Government HR shared services, NCSS=National Council of Social Service, VWO=Voluntary Welfare Organisation, IHRP=HR professional certification (Singapore), BOI=Board of Inquiry (HR/IR), NGEMR=Electronic Medical Records system, SOC=Specialist Outpatient Clinic, GCP=Good Clinical Practice, IRB=Institutional Review Board, EDC=Electronic Data Capture
   PA=People's Association (government-linked community organisation — counter service, facility management, event coordination, community ops — directly transferable to PSA and admin roles)
   SportSG / Sport Singapore=statutory board managing sports facilities and public engagement — customer service, counter operations, facility checks, event support — transferable to PSA and operations roles
   ActiveSG=Sport Singapore's membership programme — counter service and public-facing operations
   CC=Community Centre (PA-managed) — public service delivery, counter ops, admin, event management

4. JOB STABILITY — Reward long tenures and promotions:
   - Candidate promoted within the same organisation = strong performance signal, boost score
   - 5+ years at a single employer = excellent stability, especially valuable for PSA and admin roles
   - Multiple promotions across career = high performer pattern
   - Leaving shift work / rotating shifts for office hours = completely legitimate reason, do NOT treat as a red flag
   - Leaving due to contract end, restructuring, or seeking career growth = legitimate, do not penalise

4. SCORING: Base score on actual evidence in the CV. Do not penalise for information that is simply not mentioned — absence of evidence is not evidence of absence.

5. SUITABILITY: Set suitable=true if the candidate has the core competencies required, even if some nice-to-haves are missing.

6. JOB HISTORY IS A UNIVERSAL SIGNAL — applies to ALL roles, ALL industries:
   POSITIVE signals (boost score):
   - Long tenures (3+ years per employer) = loyalty, reliability, depth of experience
   - Promoted within the same organisation = strong performer, recognised by employer
   - Progressive career growth (e.g. Executive → Senior Executive → Manager) = ambition and capability
   - Consistent focus in relevant skillsets over many years = deep expertise
   NEGATIVE signals (note but do not auto-fail):
   - Job hopping: multiple roles under 12 months = flag as concern, but investigate reason first
   - Frequent unrelated industry jumps without clear progression = flag
   - Large unexplained gaps (6+ months post-2022) = worth probing, not automatic rejection
   NEUTRAL — do NOT penalise:
   - Leaving shift/rotating work for office hours = legitimate lifestyle reason
   - Contract roles ending naturally = not the candidate's choice
   - Career breaks for caregiving, health, or study = legitimate

7. RED FLAGS to note (but not automatically fail):
   - Pure support/AMS-only IT backgrounds with zero implementation experience
   - Pure bench-lab scientists applying for clinical research coordinator roles (need coordination experience)
   - Overqualified: ONLY when scope AND title are both clearly above role requirements

Respond with ONLY a valid JSON object (no markdown, no code fences, no explanation):
{{
  "candidate_name": "Full name of the candidate, or 'Unknown'",
  "score": <integer 0-100>,
  "suitable": <true or false>,
  "overqualified": <true if candidate seniority is clearly above the role, false otherwise>,
  "strengths": ["Full sentence: specific strength tied to a JD requirement with evidence from CV", "Strength 2", "Strength 3"],
  "gaps": ["Full sentence: specific gap or concern with explanation", "Gap 2"],
  "key_skills_found": ["skill1", "skill2", "skill3", "skill4", "skill5"],
  "years_experience": <number or null if unclear>,
  "requirements_assessment": [
    {{"requirement": "Full requirement text from JD", "met": true, "evidence": "Brief evidence from CV or reason not met"}},
    {{"requirement": "Requirement 2", "met": true, "evidence": "..."}},
    {{"requirement": "Requirement 3", "met": false, "evidence": "Not evidenced in CV"}}
  ],
  "email_bullets": ["Specific, factual highlight 1 a recruiter would tell a client about this candidate", "Highlight 2", "Highlight 3"]
}}

Instructions:
- strengths and gaps: write FULL sentences, no truncation, no ellipsis
- requirements_assessment: extract the 4-6 most important requirements from the JD and assess each one
- email_bullets: neutral, factual highlights — do NOT recommend for or against, just present facts
- If overqualified=true, include it as a gap: "Candidate's current seniority may be above this role — worth discussing expectations"

Scoring guide: 90-100 exceptional, 70-89 strong, 50-69 partial match (still worth considering), 30-49 weak, 0-29 poor."""

    old_handler = signal.signal(signal.SIGALRM, _alarm_handler)
    signal.alarm(25)  # hard OS interrupt — fires before Railway's 30s worker timeout
    try:
        t_api = time.time()
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1800,
            messages=[{"role": "user", "content": prompt}],
        )
        print(f"[Claude] score_resume API took {time.time()-t_api:.2f}s")
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)

    raw = message.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


@app.route("/interview-questions", methods=["POST"])
def interview_questions():
    data = request.get_json()
    job_title = (data.get("job_title") or "").strip()
    job_description = (data.get("job_description") or "").strip()
    candidate_name = data.get("candidate_name", "the candidate")
    strengths = data.get("strengths", [])
    gaps = data.get("gaps", [])
    key_skills = data.get("key_skills", [])

    if not job_title or not job_description:
        return jsonify({"error": "Job title and description are required."}), 400

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY not set."}), 500

    client = anthropic.Anthropic(api_key=api_key)

    prompt = f"""You are a senior recruiter preparing a phone screening call. Generate 5 targeted interview questions for this specific candidate and role.

Job Title: {job_title}
Job Description:
{job_description[:3000]}

Candidate: {candidate_name}
Their strengths: {', '.join(strengths[:3])}
Their gaps: {', '.join(gaps[:3])}
Key skills found: {', '.join(key_skills[:5])}

Generate 5 role-specific questions that:
1. Probe the candidate's actual experience relevant to THIS job
2. Dig into any gaps or areas needing verification
3. Are open-ended and encourage detailed answers
4. Are appropriate for a phone/video screening call (not a deep technical interview)

Respond with ONLY a valid JSON object (no markdown, no explanation):
{{
  "questions": [
    {{
      "category": "Short category label e.g. 'Leadership' or 'Technical' or 'Industry knowledge'",
      "question": "Full question text",
      "why": "One sentence: why this question is important for this role/candidate"
    }}
  ]
}}"""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return jsonify(json.loads(raw))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    return render_template("index.html")



@app.route("/debug-env")
def debug_env():
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    return jsonify({
        "key_found": bool(key),
        "key_preview": key[:12] + "..." if key else "NOT SET",
        "env_vars": [k for k in os.environ.keys()]
    })


@app.route("/analyze-jd", methods=["POST"])
def analyze_jd():
    data = request.get_json()
    job_title = (data.get("job_title") or "").strip()
    job_description = (data.get("job_description") or "").strip()

    if not job_title or not job_description:
        return jsonify({"error": "Job title and job description are required."}), 400

    try:
        result = analyze_jd_with_claude(job_title, job_description)
        return jsonify(result)
    except json.JSONDecodeError:
        return jsonify({"error": "Failed to parse JD analysis response."}), 500
    except anthropic.APIError as e:
        return jsonify({"error": f"API error — {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/extract-jd", methods=["POST"])
def extract_jd():
    """Extract text from an uploaded JD file (PDF or DOCX)."""
    file = request.files.get("jd_file")
    if not file or file.filename == "":
        return jsonify({"error": "No file uploaded."}), 400

    filename = secure_filename(file.filename)
    if not allowed_file(filename):
        return jsonify({"error": "Unsupported file type. Use PDF or DOCX."}), 400

    try:
        file_bytes = file.read()
        text = extract_text(file_bytes, filename)
        if not text:
            return jsonify({"error": "Could not extract text from file."}), 400
        return jsonify({"text": text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/screen", methods=["POST"])
def screen():
    job_title = request.form.get("job_title", "").strip()
    job_description = request.form.get("job_description", "").strip()

    if not job_title or not job_description:
        return jsonify({"error": "Job title and job description are required."}), 400

    files = request.files.getlist("resumes")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "Please upload at least one resume."}), 400

    results = []
    errors = []

    # Fetch past placed examples for this role to guide scoring
    examples = get_placed_examples(job_title)

    for file in files:
        if file.filename == "":
            continue
        filename = secure_filename(file.filename)
        if not allowed_file(filename):
            errors.append(f"{filename}: unsupported file type (use PDF or DOCX).")
            continue

        try:
            file_bytes = file.read()
            cv_text = extract_text(file_bytes, filename)

            if not cv_text:
                errors.append(f"{filename}: could not extract text from file.")
                continue

            result = score_resume_with_claude(cv_text, job_title, job_description, examples=examples)
            result["filename"] = filename
            results.append(result)

        except _ClaudeTimeout:
            print(f"[Claude] score_resume TIMED OUT for {filename}")
            errors.append(f"{filename}: screening timed out — upload this file on its own and try again.")
        except json.JSONDecodeError:
            errors.append(f"{filename}: Claude returned an unexpected response format.")
        except anthropic.APIError as e:
            errors.append(f"{filename}: API error — {str(e)}")
        except Exception as e:
            errors.append(f"{filename}: {str(e)}")

    results.sort(key=lambda x: x.get("score", 0), reverse=True)
    return jsonify({"results": results, "errors": errors})


@app.route("/export", methods=["POST"])
def export():
    data = request.get_json()
    results = data.get("results", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    headers = ["Rank", "Candidate Name", "Score", "Suitable", "Key Skills", "Strengths", "Gaps", "Email Bullets"]
    ws.append(headers)

    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1D9E75")
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for rank, r in enumerate(results, 1):
        score = r.get("score", 0)
        suitable = "Yes" if r.get("suitable") else "No"
        skills = ", ".join(r.get("key_skills_found", []))
        strengths = "\n".join(f"• {s}" for s in r.get("strengths", []))
        gaps = "\n".join(f"• {g}" for g in r.get("gaps", []))
        email_bullets = "\n".join(f"• {b}" for b in r.get("email_bullets", []))

        row = [rank, r.get("candidate_name", "Unknown"), score, suitable, skills, strengths, gaps, email_bullets]
        ws.append(row)

        score_cell = ws.cell(row=rank + 1, column=3)
        if score >= 90:
            score_cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="16A34A")
        elif score >= 70:
            score_cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="DCFCE7")
        elif score >= 50:
            score_cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FEF9C3")
        elif score >= 30:
            score_cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FED7AA")
        else:
            score_cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FEE2E2")

        for col in [6, 7, 8]:
            ws.cell(row=rank + 1, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    col_widths = [6, 22, 8, 10, 32, 40, 32, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    for row_num in range(2, len(results) + 2):
        ws.row_dimensions[row_num].height = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resume_screening_results.xlsx",
    )


@app.route("/pipeline", methods=["GET"])
def get_pipeline():
    conn = get_db()
    try:
        rows = conn.execute('SELECT * FROM pipeline ORDER BY updated_at DESC').fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


@app.route("/pipeline/save", methods=["POST"])
def save_pipeline():
    data = request.get_json()
    try:
        conn = get_db()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    try:
        now = datetime.utcnow().isoformat()
        existing = conn.execute(
            'SELECT id FROM pipeline WHERE candidate_name=? AND role_title=?',
            (data.get('candidate_name',''), data.get('role_title',''))
        ).fetchone()
        if existing:
            conn.execute('UPDATE pipeline SET stage=?, updated_at=? WHERE id=?',
                         (data.get('stage','submitted'), now, existing['id']))
            pipeline_id = existing['id']
        else:
            cur = conn.execute(
                '''INSERT INTO pipeline
                   (candidate_name,filename,role_title,score,stage,key_skills,strengths,years_experience,created_at,updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)''',
                (data.get('candidate_name','Unknown'), data.get('filename',''),
                 data.get('role_title',''), data.get('score',0),
                 data.get('stage','submitted'),
                 json.dumps(data.get('key_skills',[])),
                 json.dumps(data.get('strengths',[])),
                 data.get('years_experience'), now, now)
            )
            pipeline_id = cur.lastrowid
        conn.commit()
        return jsonify({'id': pipeline_id, 'status': 'saved'})
    finally:
        conn.close()


@app.route("/pipeline/update", methods=["POST"])
def update_pipeline():
    data = request.get_json()
    conn = get_db()
    try:
        now = datetime.utcnow().isoformat()
        conn.execute('UPDATE pipeline SET stage=?, updated_at=? WHERE id=?',
                     (data.get('stage'), now, data.get('id')))
        conn.commit()
        return jsonify({'status': 'updated'})
    finally:
        conn.close()


if __name__ == "__main__":
    app.run(debug=True)
